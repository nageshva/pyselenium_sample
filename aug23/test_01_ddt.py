import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

def get_test_data():
    workbook=load_workbook("testdata2.xlsx")
    sheet=workbook.active
    data =[]
    for row in sheet.iter_rows(min_row=2,values_only=True):
        data.append(row)
    return data


@pytest.fixture
def driver():
    driver=webdriver.Chrome()
    driver.get("https://app.vwo.com/")
    driver.maximize_window()
    yield driver
    driver.quit()

@pytest.mark.parametrize("username,password,result",get_test_data())
def test_vwo_login(driver,username,password,result):
    email_ele = driver.find_element(By.ID, "login-username")
    email_ele.send_keys(username)
    pasword_ele = driver.find_element(By.ID, "login-password")
    pasword_ele.send_keys(password)

    driver.find_element(By.ID, "js-login-btn").click()
    time.sleep(3)

    print(username, password,driver.current_url)
    if result == "fail":
        error_msg = driver.find_element(By.ID, "js-notification-box-msg").text
        assert error_msg in "Your email, password, IP address or location did not match"
    else:
        assert "https://app.vwo.com/#/dashboard" in driver.current_url

