import openpyxl

import time

from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By

def getRowcount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColcount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.cell(rownum,colnum).value)

def writeData(file,sheetName,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(rownum,colnum).value=data
    workbook.save(file)


def fillGreenColor(file,sheetName,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    greenfill=PatternFill(start_color='60b212',
                          end_color='60b212'
                                    ,fill_type='solid')
    sheet.cell(rownum,colnum).fill=greenfill
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill = PatternFill(start_color='ff0000',
                            end_color='ff0000'
                            , fill_type='solid')
    sheet.cell(rownum, colnum).fill = redfill
    workbook.save(file)






