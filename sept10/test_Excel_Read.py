import openpyxl

import time

from selenium import webdriver
from selenium.webdriver.common.by import By

def test_readExcel():

#file ----> Workbook---> Sheets---->rows/columns[cells]

    #file="C:\\Users\\yours\\Desktop\\Test.xlsx"
    file="C:\\Users\\yours\\Desktop\\New.xlsx"
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.active
    #
    # rows=sheet.max_row
    # cols=sheet.max_column
    #
    # for r in range (2,rows):
    #     for c in range(1,cols+1):
    #         sheet.cell(r,c).value="Hello" # write data on to excel
          #  print( sheet.cell(r,c).value)  read data from excel

    sheet.cell(1,1).value ="Hello"
    sheet.cell(1,2).value ="world"

    workbook.save(file)

        # print()
